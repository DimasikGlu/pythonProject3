# 1. Установите openpyxl в свой проект
from openpyxl import *
from openpyxl.styles import *

# 2. Прочтите файл excel
wb1 = load_workbook('ExcelForPython.xlsx')
# 3. Из файла прочтите «первый лист» данных
print(wb1.sheetnames)
wb1_values = wb1.worksheets[0]

# 4. Получите значение F19
# wb4_values = wb1_values['F19'].value
# print(wb4_values)
# 5. Получите значение в диапазоне [‘A18:E30’] и назначьте ее в
# значение myListExcel
# myListExcel = wb1_values['A18:E30']
# 6. Сделайте перебор всех данных с этого myListExcel
# c помощью for цикла
# for row in myListExcel:
#         for i in row:
#                 print(i.value, end=" ")
#         print()
# 7. Сделайте перебор данных установив лимит в 5 рядов из myListExcel
# myListExcel2 = myListExcel.cell(row=5,colum=5).value
# print(myListExcel2)
# for i in myListExcel2:
#         for cell in i:
#                 print(cell.value, end="")
#         print()


# 8. Попросить вести пользователя следующие данные, как в след.
# примере:
# [‘Имя’,’Возраст’,’Зарплата’,’Год работ’],
# [‘Самат’, ‘Улан’]
# [25, 30]
# [25000, 32000]
# [5,7]
from openpyxl import workbook
# wb2 = Workbook('myListExcel2')
# myListExcel2 = wb2.create_sheet("MySheet")
# for i in range(1):
#         myListExcel2.append({1:'Plese input Имя'})

# 9. Введенные данные от пользователя необходимо записать в отдельный
# файл Excel с названием данные о работниках
# «workers.xlsx»
# Пример данных в таблице Excel:

#
# ({'Имя': input("Plese input Имя:"),'Возраст': input("Plese input возраст:"),'Зарплата': input("Plese input Зарплата:"),
#                      'Год работ': input("Plese input Год работ:")})